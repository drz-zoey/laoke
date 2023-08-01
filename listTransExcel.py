import sys
import re
import openpyxl

# 打开excel表格，新建或打开已有表格
workbook = openpyxl.Workbook()
worksheet = workbook.active

path = "C:\\Users\\John\\Desktop\\laoke"
#path = "F:\\Material\\互助\\code\\laoke"

input_file = path + "\\example.txt"
output_file = path + "\\result.xlsx"

#input_file = sys.argv[1]
#output_file = sys.argv[2] + "\\result.xlsx"


# 逐行读取文本文件
with open(input_file, 'r' , encoding='utf-8') as f:
    line_num = 2
    undo_num = 3
    worksheet.cell(row=line_num, column=2).value = '品牌'
    worksheet.cell(row=line_num, column=3).value = '商品'
    worksheet.cell(row=line_num, column=4).value = '价格（元）'
    worksheet.cell(row=line_num, column=5).value = '来源'
    worksheet.cell(row=line_num, column=7).value = '待处理'
    line_num = line_num+1

    for line in f:
        # 提取每行末尾的数字，作为价格
        price_match = re.search(r"\d+$", line.strip())
        if price_match:
            price = price_match.group()
            name = line[:price_match.start()].strip()
        else:
            price = ""
            name = line.strip()

        # 提取出现的品牌名称
        brand_match = re.search(r"(兰蔻|雅诗兰黛|卡地亚|圣罗兰|蒂佳婷|cpb肌肤之钥|海蓝之谜|娇韵诗|阿玛尼|赫莲娜|科颜氏|Gucci|倩碧|黛珂|资生堂|馥蕾诗|欧莱雅|迪奥|莱珀妮|赫莲娜|娇兰|SK-II|植村秀|Tom Ford|NARS|MAC|魅可|whoo后|雅萌|ahc|飞利浦|百瑞德|兰芝|香奈儿|银座|衰败城市|伊菲丹|芭比布朗|祖玛龙|希思黎|碧欧泉|欧舒丹|修丽可|爱马仕|蔻依|祖玛珑|梅森马吉拉)", line)
       
        #根据单品名称匹配品牌
        brand_match_skii = re.search(r"(SK2|sk2|skii|sk-ii|眼面套|神灯套|小灯泡|神仙水|大红瓶)", line)
        brand_match_ysld = re.search(r"(增量五件套|抗衰|红石榴|小棕瓶|抗蓝光|眼绷带|白金眼霜|白金粉底液|沁水|dw|原生液|樱花水|智妍|抗衰老|黑松露|白金面霜)", line)
        brand_match_hlzm = re.search(r"(LAMER|Lamer|lamer|精粹水|精萃乳|海南之谜|舒缓喷雾|浓缩精华|经典面霜)", line)
        brand_match_zst = re.search(r"(安耐晒|安热沙|蓝胖子|红腰子|悦薇)", line)
        brand_match_cpb = re.search(r"(cpb|肌肤之钥|CPB|Cpb)", line)
        brand_match_whoo = re.search(r"(whoo)", line)
        brand_match_tf = re.search(r"(TF|Tf|tf|tom ford|Tom Ford|TOM FORD|乌木)", line)
        brand_match_xsl = re.search(r"(希思黎|sisley|Sisley|全能乳)", line)
        brand_match_xlk = re.search(r"(修丽可|色修|紫米精华|CE精华|SCF抗氧)", line)
        brand_match_ahc = re.search(r"(ahc|Ahc|AHC|黄金面膜)", line)
        brand_match_nars = re.search(r"(nars|Nars|纳斯|超方瓶|大白饼)", line)
        brand_match_msmjl = re.search(r"(马吉拉|梅森马吉拉)", line)
        brand_match_lk = re.search(r"(小白管|菁纯|小黑瓶|发光眼霜|雪花霜|粉水|极光)", line)
        brand_match_kys = re.search(r"(金盏花|宝宝霜|高保湿|白泥|紫波A面霜)", line)
        brand_match_qb = re.search(r"(倩碧|紫光|镭射)", line)
        brand_match_amn = re.search(r"(玉龙茶香|黑曜石|黑钥匙|myway|MY WAY|My Way|my way|自我无界|大师|权力)", line)
        brand_match_ysl = re.search(r"(圣罗兰|YSL|ysl|Ysl|夜皇后|藏金)", line)
        brand_match_gucci = re.search(r"(Gucci|gucci|古驰)", line)
        brand_match_hr = re.search(r"(Hr|HR|赫莲娜|黑绷带|白绷带|绿宝瓶|黑白绷带|小针管)", line)
        brand_match_sbcs = re.search(r"(衰败城市|牛郎|URBAN DECAY|Urban Decay|urban decay)", line)
        brand_match_ky = re.search(r"(寇依|蔻依|chole|chloe|Chloe|CHLOE|北国雪松|木兰诗语)", line)
        #brand_match_zml = re.search(r"(祖玛珑|祖马龙|蓝风铃|)", line)
        brand_match_phlg = re.search(r"(PENHALIGON'S|潘海利根|兽首)", line)
        brand_match_pola = re.search(r"(POLA|pola|Pola|宝丽|黑BA|黑ba|黑Ba)", line)
        
        if brand_match:
            brand = brand_match.group()
        elif brand_match_skii:
            brand = "SK-II"
        elif brand_match_ysld:
            brand = "雅诗兰黛"
        elif brand_match_hlzm:
            brand = "海蓝之谜"
        elif brand_match_zst:
            brand = "资生堂"
        elif brand_match_cpb:
            brand = "cpb肌肤之钥"
        elif brand_match_whoo:
            brand = "whoo后"
        elif brand_match_tf:
            brand = "Tom Ford"
        elif brand_match_xsl:
            brand = "Sisley希思黎"
        elif brand_match_xlk:
            brand = "修丽可"
        elif brand_match_ahc:
            brand = "AHC爱和纯"
        elif brand_match_nars:
            brand = "NARS"
        elif brand_match_msmjl:
            brand = "梅森马吉拉"
        elif brand_match_lk:
            brand = "兰蔻"
        elif brand_match_kys:
            brand = "科颜氏"
        elif brand_match_qb:
            brand = "倩碧"
        elif brand_match_amn:
            brand = "Amani阿玛尼"
        elif brand_match_ysl:
            brand = "Ysl圣罗兰"
        elif brand_match_gucci:
            brand = "Gucci"
        elif brand_match_hr:
            brand = "赫莲娜"
        elif brand_match_ky:
            brand = "chloe蔻依"
        elif brand_match_phlg:
            brand = "潘海利根"
        elif brand_match_pola:
            brand = "POLA 宝丽"
        else:
            brand = ""
            

        # 提取括号内的内容，作为来源地
        source_match = re.search(r"\（(.*?)\）", line)
        if source_match:
            source = source_match.group(1)
        else:
            source = ""

        print("brand:"+brand)
        print("name:"+name)
        print("price:"+price)
        print("source:"+source)
        
        # 将提取到的信息及匹配失败项写入excel表格
        if price or brand or source:
            worksheet.cell(line_num, 2, brand)
            worksheet.cell(line_num, 3, name)
            worksheet.cell(line_num, 4, price)
            worksheet.cell(line_num, 5, source)
            line_num = line_num + 1
        else:
            undo = line.strip()
            if undo:
                worksheet.cell(undo_num, 7, line.strip())
                undo_num = undo_num + 1

# 保存excel表格
workbook.save(output_file)